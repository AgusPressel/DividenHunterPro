"""
================================================================================
MÓDULO 3: INTERFAZ DE USUARIO CON STREAMLIT (Frontend)
================================================================================

CONCEPTO TEÓRICO:
-----------------
Streamlit es un framework de Python que permite crear aplicaciones web
interactivas SIN necesidad de HTML/CSS/JavaScript.

VENTAJAS PARA FINANCIAS:
- Desarrollo rápido (ideal para prototipos)
- Integración nativa con pandas y plotly
- Componentes interactivos out-of-the-box
- Perfecto para dashboards financieros

ARQUITECTURA STREAMLIT:
- app.py es el punto de entrada
- Cada vez que el usuario interactúa, Streamlit re-ejecuta el script
- Usamos st.session_state para mantener estado entre interacciones

IMPORTADOR EXCEL:
-----------------
En fintech, es común recibir listas de tickers en Excel. Este módulo
demuestra cómo procesar archivos Excel y usar la lógica del Módulo 1
para analizar múltiples activos en batch.
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
from typing import List, Dict
import sys
import os

# Importar módulos anteriores (simulando que están en el mismo directorio)
# En producción, estos serían imports normales
sys.path.append(os.path.dirname(__file__))

from modulo1_ingenieria_datos import DividendAnalyzer
from modulo2_persistencia_datos import DatabaseManager


class StreamlitApp:
    """
    Clase que encapsula la lógica de la interfaz Streamlit.
    
    PRINCIPIO DE DISEÑO: Separación de UI y Lógica
    - Esta clase orquesta la UI
    - Delega la lógica de negocio a los otros módulos
    """
    
    def __init__(self):
        """Inicializa la aplicación."""
        self.analyzer = DividendAnalyzer()
        self.db = DatabaseManager()
        self._setup_page()
    
    def _setup_page(self):
        """
        Configura la página de Streamlit.
        
        Este método establece el título, icono, y layout general.
        """
        st.set_page_config(
            page_title="Dividend Hunter Pro",
            page_icon="💰",
            layout="wide",
            initial_sidebar_state="expanded"
        )
        
        # CSS personalizado para mejorar la UI
        st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            font-weight: bold;
            color: #1f77b4;
            text-align: center;
            margin-bottom: 2rem;
        }
        .metric-card {
            background-color: #f0f2f6;
            padding: 1rem;
            border-radius: 0.5rem;
            margin: 0.5rem 0;
        }
        </style>
        """, unsafe_allow_html=True)
    
    def render_header(self):
        """Renderiza el encabezado principal."""
        st.markdown('<h1 class="main-header">💰 Dividend Hunter Pro</h1>', 
                   unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("""
        **Bienvenido a Dividend Hunter Pro**
        
        Esta aplicación te permite:
        - 🔍 Buscar activos y analizar sus dividendos
        - 📊 Importar listas de tickers desde Excel
        - 📈 Visualizar métricas financieras clave
        - 💎 Encontrar "gemas" (alto yield, bajo costo)
        """)
    
    def render_sidebar(self):
        """
        Renderiza el sidebar con opciones de navegación.
        
        El sidebar es el "centro de control" de la aplicación.
        """
        st.sidebar.title("🎯 Navegación")
        
        page = st.sidebar.radio(
            "Selecciona una opción:",
            ["🏠 Inicio", "📥 Importar Excel", "🔍 Buscar Activo", 
             "📊 Ver Activos", "📈 Estadísticas"]
        )
        
        st.sidebar.markdown("---")
        st.sidebar.markdown("### ℹ️ Información")
        st.sidebar.info("""
        **Dividend Hunter Pro** analiza automáticamente la frecuencia
        de pago de dividendos analizando el historial de 12 meses.
        """)
        
        return page
    
    def process_excel_file(self, uploaded_file) -> List[str]:
        """
        FUNCIÓN CLAVE: Procesa archivo Excel y extrae tickers.
        
        Esta función demuestra cómo usar openpyxl para leer Excel
        y extraer símbolos de tickers para procesamiento batch.
        
        Args:
            uploaded_file: Archivo subido a Streamlit
        
        Returns:
            Lista de símbolos de tickers
        """
        try:
            # Leer el archivo Excel
            # Streamlit proporciona el archivo como BytesIO
            wb = load_workbook(io.BytesIO(uploaded_file.read()))
            
            # Obtener la primera hoja
            ws = wb.active
            
            tickers = []
            
            # Estrategia: Buscar tickers en la primera columna
            # (Puedes adaptar esto según tu formato de Excel)
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0]:  # Si la primera celda tiene contenido
                    ticker = str(row[0]).strip().upper()
                    # Validar que parece un ticker (letras y números, 1-5 caracteres)
                    if ticker.isalnum() and 1 <= len(ticker) <= 5:
                        tickers.append(ticker)
            
            return list(set(tickers))  # Eliminar duplicados
            
        except Exception as e:
            st.error(f"❌ Error procesando Excel: {e}")
            return []
    
    def import_excel_page(self):
        """
        Página para importar y procesar archivos Excel.
        
        Esta es la funcionalidad principal del módulo: permite subir
        un Excel con tickers y procesarlos usando la lógica del Módulo 1.
        """
        st.header("📥 Importar Activos desde Excel")
        
        st.markdown("""
        **Instrucciones:**
        1. Prepara un archivo Excel con los símbolos de tickers en la primera columna
        2. Sube el archivo usando el botón de abajo
        3. La aplicación analizará cada ticker y guardará los resultados
        """)
        
        uploaded_file = st.file_uploader(
            "Selecciona archivo Excel (.xlsx)",
            type=['xlsx'],
            help="El archivo debe tener los tickers en la primera columna"
        )
        
        if uploaded_file is not None:
            with st.spinner("Procesando archivo Excel..."):
                tickers = self.process_excel_file(uploaded_file)
                
                if not tickers:
                    st.warning("⚠️ No se encontraron tickers válidos en el archivo")
                    return
                
                st.success(f"✅ Se encontraron {len(tickers)} tickers únicos")
                st.write("**Tickers encontrados:**", ", ".join(tickers))
                
                # Procesar cada ticker
                if st.button("🚀 Analizar y Guardar Activos", type="primary"):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    success_count = 0
                    error_count = 0
                    
                    for i, ticker in enumerate(tickers):
                        status_text.text(f"Analizando {ticker}... ({i+1}/{len(tickers)})")
                        
                        # Usar la lógica del Módulo 1
                        metrics = self.analyzer.get_asset_metrics(ticker)
                        
                        if metrics:
                            # Guardar en BD usando Módulo 2
                            if self.db.upsert_asset(metrics):
                                success_count += 1
                            else:
                                error_count += 1
                        else:
                            error_count += 1
                        
                        progress_bar.progress((i + 1) / len(tickers))
                    
                    status_text.empty()
                    progress_bar.empty()
                    
                    st.success(f"✅ Procesados: {success_count} exitosos, {error_count} con errores")
    
    def search_asset_page(self):
        """Página para buscar un activo individual."""
        st.header("🔍 Buscar Activo")
        
        symbol = st.text_input(
            "Ingresa el símbolo del activo (ej: AAPL, MSFT, O)",
            value="",
            help="Usa el símbolo de ticker estándar (NYSE, NASDAQ, etc.)"
        ).upper()
        
        if symbol:
            if st.button("🔍 Buscar", type="primary"):
                with st.spinner(f"Analizando {symbol}..."):
                    metrics = self.analyzer.get_asset_metrics(symbol)
                    
                    if metrics:
                        # Mostrar resultados
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            st.metric("Precio", f"${metrics['current_price']:.2f}")
                        with col2:
                            st.metric("Dividend Yield", f"{metrics['dividend_yield']:.2f}%")
                        with col3:
                            st.metric("Dividendo Anual", f"${metrics['annual_dividend']:.2f}")
                        with col4:
                            freq_emoji = {
                                'mensual': '📅',
                                'trimestral': '📆',
                                'irregular': '⚠️',
                                'sin_dividendos': '❌'
                            }
                            emoji = freq_emoji.get(metrics['dividend_frequency'], '❓')
                            st.metric("Frecuencia", 
                                    f"{emoji} {metrics['dividend_frequency'].upper()}")
                        
                        # Información adicional
                        st.markdown("### 📋 Detalles")
                        st.json(metrics)
                        
                        # Botón para guardar
                        if st.button("💾 Guardar en Base de Datos"):
                            if self.db.upsert_asset(metrics):
                                st.success(f"✅ {symbol} guardado correctamente")
                            else:
                                st.error("❌ Error al guardar")
                    else:
                        st.error(f"❌ No se pudieron obtener datos para {symbol}")
    
    def view_assets_page(self):
        """Página para ver todos los activos guardados."""
        st.header("📊 Activos Guardados")
        
        # Filtro por frecuencia
        filter_freq = st.selectbox(
            "Filtrar por frecuencia:",
            ["Todos", "mensual", "trimestral", "irregular", "sin_dividendos"]
        )
        
        freq_filter = None if filter_freq == "Todos" else filter_freq
        
        # Obtener activos
        assets = self.db.get_all_assets(freq_filter)
        
        if assets:
            # Convertir a DataFrame para mejor visualización
            df = pd.DataFrame(assets)
            
            # Mostrar tabla
            st.dataframe(
                df[['symbol', 'name', 'current_price', 'dividend_yield', 
                    'dividend_frequency', 'last_updated']],
                use_container_width=True
            )
            
            st.info(f"📊 Total: {len(assets)} activos")
        else:
            st.warning("⚠️ No hay activos guardados. Usa 'Importar Excel' o 'Buscar Activo'")
    
    def stats_page(self):
        """Página de estadísticas."""
        st.header("📈 Estadísticas")
        
        stats = self.db.get_stats()
        
        if stats:
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Activos", stats.get('total_assets', 0))
            
            with col2:
                st.metric("Yield Promedio", f"{stats.get('average_yield', 0)}%")
            
            with col3:
                freq_dist = stats.get('frequency_distribution', {})
                st.metric("Frecuencias", len(freq_dist))
            
            # Gráfico de distribución
            if freq_dist:
                st.markdown("### Distribución por Frecuencia")
                freq_df = pd.DataFrame(
                    list(freq_dist.items()),
                    columns=['Frecuencia', 'Cantidad']
                )
                st.bar_chart(freq_df.set_index('Frecuencia'))
    
    def run(self):
        """
        Método principal que ejecuta la aplicación.
        
        Este es el "orquestador" que decide qué página mostrar
        según la selección del usuario.
        """
        self.render_header()
        page = self.render_sidebar()
        
        if page == "🏠 Inicio":
            st.info("👈 Selecciona una opción en el sidebar para comenzar")
        
        elif page == "📥 Importar Excel":
            self.import_excel_page()
        
        elif page == "🔍 Buscar Activo":
            self.search_asset_page()
        
        elif page == "📊 Ver Activos":
            self.view_assets_page()
        
        elif page == "📈 Estadísticas":
            self.stats_page()


# ============================================================================
# PUNTO DE ENTRADA PARA STREAMLIT
# ============================================================================

if __name__ == "__main__":
    """
    Para ejecutar esta aplicación:
    
    streamlit run modulo3_interfaz_usuario.py
    """
    app = StreamlitApp()
    app.run()

