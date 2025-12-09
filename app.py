"""
================================================================================
DIVIDEND HUNTER PRO - APLICACIÓN COMPLETA
================================================================================

Este es el archivo principal que unifica todos los módulos en una aplicación
completa y lista para producción.

ARQUITECTURA:
-------------
- Módulo 1: Ingeniería de Datos (DividendAnalyzer)
- Módulo 2: Persistencia (DatabaseManager)
- Módulo 3: Interfaz (StreamlitApp)
- Módulo 4: Visualización (FinancialVisualizer)
- Módulo 5: Utilidades (ErrorHandler, Config, etc.)

EJECUCIÓN:
----------
streamlit run app.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
from typing import List, Dict, Optional
import sys
import os
import logging

# Importar todos los módulos
from modulo1_ingenieria_datos import DividendAnalyzer
from modulo2_persistencia_datos import DatabaseManager
from modulo4_visualizacion import FinancialVisualizer
from modulo5_refactorizacion import (
    ErrorHandler, Config, DataValidator, 
    format_currency, format_percentage
)

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DividendHunterApp:
    """
    Clase principal que orquesta toda la aplicación.
    
    Esta clase integra todos los módulos en una aplicación cohesiva.
    """
    
    def __init__(self):
        """Inicializa todos los componentes de la aplicación."""
        try:
            self.analyzer = DividendAnalyzer()
            # Usar la misma ruta de BD para todas las instancias
            self.db_path = Config.DB_PATH
            self.db = DatabaseManager(self.db_path)
            # Pasar la misma ruta de BD al visualizador
            self.visualizer = FinancialVisualizer(self.db_path)
            self._setup_page()
            logger.info(f"Aplicación inicializada correctamente. BD: {self.db_path}")
        except Exception as e:
            logger.error(f"Error inicializando aplicación: {e}", exc_info=True)
            st.error("❌ Error inicializando la aplicación. Por favor, revisa los logs.")
            raise
    
    def _setup_page(self):
        """Configura la página de Streamlit."""
        st.set_page_config(
            page_title=Config.PAGE_TITLE,
            page_icon=Config.PAGE_ICON,
            layout="wide",
            initial_sidebar_state="expanded"
        )
        
        # CSS personalizado
        st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            font-weight: bold;
            color: #1f77b4;
            text-align: center;
            margin-bottom: 2rem;
        }
        /* Asegurar que las métricas de Streamlit se muestren correctamente */
        [data-testid="stMetricValue"] {
            visibility: visible !important;
            opacity: 1 !important;
        }
        [data-testid="stMetricLabel"] {
            visibility: visible !important;
            opacity: 1 !important;
        }
        </style>
        """, unsafe_allow_html=True)
    
    def render_header(self):
        """Renderiza el encabezado principal."""
        st.markdown(f'<h1 class="main-header">{Config.PAGE_ICON} {Config.PAGE_TITLE}</h1>', 
                   unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("""
        **Bienvenido a Dividend Hunter Pro**
        
        Esta aplicación te permite:
        - 🔍 Buscar activos y analizar sus dividendos automáticamente
        - 📥 Importar listas de tickers desde Excel
        - 📊 Visualizar métricas financieras clave
        - 💎 Encontrar "gemas" de inversión (alto yield, bajo costo)
        - 📈 Analizar la frecuencia de pago de dividendos (Mensual vs Trimestral)
        """)
    
    def render_sidebar(self):
        """Renderiza el sidebar con navegación."""
        st.sidebar.title("🎯 Navegación")
        
        page = st.sidebar.radio(
            "Selecciona una opción:",
            ["🏠 Inicio", "📥 Importar Excel", "🔍 Buscar Activo", 
             "📊 Ver Activos", "📈 Visualizaciones", "ℹ️ Estadísticas", 
             "🔧 Mantenimiento", "🏪 Gestión de Plataformas", "💼 Constructor de Portfolio"]
        )
        
        st.sidebar.markdown("---")
        st.sidebar.markdown("### ℹ️ Información")
        st.sidebar.info("""
        **Dividend Hunter Pro** analiza automáticamente la frecuencia
        de pago de dividendos analizando el historial de 12 meses.
        
        **Frecuencias detectadas:**
        - 📅 Mensual: 10-12 pagos/año
        - 📆 Trimestral: 3-4 pagos/año
        - ⚠️ Irregular: 1-2 pagos/año
        """)
        
        return page
    
    @ErrorHandler.handle_api_error
    def process_excel_file(self, uploaded_file) -> List[str]:
        """Procesa archivo Excel y extrae tickers."""
        try:
            wb = load_workbook(io.BytesIO(uploaded_file.read()))
            ws = wb.active
            
            tickers = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0]:
                    ticker = str(row[0]).strip().upper()
                    if ErrorHandler.validate_ticker_symbol(ticker):
                        tickers.append(ticker)
            
            return list(set(tickers))
        except Exception as e:
            logger.error(f"Error procesando Excel: {e}")
            st.error(f"❌ Error procesando archivo: {e}")
            return []
    
    def import_excel_page(self):
        """Página para importar Excel."""
        st.header("📥 Importar Activos desde Excel")
        
        st.markdown("""
        **Instrucciones:**
        1. Prepara un archivo Excel (.xlsx) con los símbolos de tickers en la primera columna
        2. Sube el archivo usando el botón de abajo
        3. La aplicación analizará cada ticker automáticamente
        4. Los resultados se guardarán en la base de datos
        """)
        
        uploaded_file = st.file_uploader(
            "Selecciona archivo Excel (.xlsx)",
            type=['xlsx'],
            help="El archivo debe tener los tickers en la primera columna"
        )
        
        if uploaded_file is not None:
            with st.spinner("Procesando archivo Excel..."):
                tickers = self.process_excel_file(uploaded_file)
                
                if not tickers:
                    st.warning("⚠️ No se encontraron tickers válidos en el archivo")
                    return
                
                st.success(f"✅ Se encontraron {len(tickers)} tickers únicos")
                
                # Mostrar tickers encontrados
                with st.expander("Ver tickers encontrados"):
                    st.write(", ".join(tickers))
                
                # Procesar tickers
                if st.button("🚀 Analizar y Guardar Activos", type="primary"):
                    self._process_tickers_batch(tickers)
    
    def _process_tickers_batch(self, tickers: List[str]):
        """Procesa un lote de tickers."""
        progress_bar = st.progress(0)
        status_text = st.empty()
        results_container = st.container()
        
        success_count = 0
        error_count = 0
        results = []
        
        for i, ticker in enumerate(tickers):
            status_text.text(f"Analizando {ticker}... ({i+1}/{len(tickers)})")
            
            try:
                metrics = self.analyzer.get_asset_metrics(ticker)
                
                if not metrics:
                    error_count += 1
                    logger.warning(f"No se obtuvieron métricas para {ticker}")
                    results.append({
                        'Ticker': ticker,
                        'Estado': '❌ Sin datos (API)',
                        'Nombre': 'N/A',
                        'Yield': 'N/A',
                        'Frecuencia': 'N/A'
                    })
                elif not DataValidator.validate_asset_metrics(metrics):
                    error_count += 1
                    logger.warning(f"Métricas inválidas para {ticker}: {metrics}")
                    results.append({
                        'Ticker': ticker,
                        'Estado': '❌ Datos inválidos',
                        'Nombre': metrics.get('name', 'N/A'),
                        'Yield': f"{metrics.get('dividend_yield', 0):.2f}%",
                        'Frecuencia': 'N/A'
                    })
                else:
                    # Intentar guardar
                    logger.info(f"Guardando {ticker} en BD...")
                    if self.db.upsert_asset(metrics):
                        success_count += 1
                        logger.info(f"✅ {ticker} guardado exitosamente")
                        results.append({
                            'Ticker': ticker,
                            'Estado': '✅ Exitoso',
                            'Nombre': metrics.get('name', 'N/A'),
                            'Yield': f"{metrics.get('dividend_yield', 0):.2f}%",
                            'Frecuencia': metrics.get('dividend_frequency', 'N/A')
                        })
                    else:
                        error_count += 1
                        logger.error(f"❌ Error al guardar {ticker} en BD")
                        results.append({
                            'Ticker': ticker,
                            'Estado': '❌ Error BD',
                            'Nombre': metrics.get('name', 'N/A'),
                            'Yield': f"{metrics.get('dividend_yield', 0):.2f}%",
                            'Frecuencia': metrics.get('dividend_frequency', 'N/A')
                        })
            except Exception as e:
                logger.error(f"Error procesando {ticker}: {e}")
                error_count += 1
                results.append({
                    'Ticker': ticker,
                    'Estado': f'❌ Error: {str(e)[:30]}',
                    'Nombre': 'N/A',
                    'Yield': 'N/A',
                    'Frecuencia': 'N/A'
                })
            
            progress_bar.progress((i + 1) / len(tickers))
        
        status_text.empty()
        progress_bar.empty()
        
        # Mostrar resumen
        st.success(f"✅ Procesamiento completado: {success_count} exitosos, {error_count} con errores")
        
        # Mostrar tabla de resultados
        if results:
            results_df = pd.DataFrame(results)
            st.dataframe(results_df, use_container_width=True)
    
    def search_asset_page(self):
        """Página para buscar un activo."""
        st.header("🔍 Buscar Activo")
        
        # Inicializar session_state si no existe
        if 'last_searched_symbol' not in st.session_state:
            st.session_state.last_searched_symbol = None
        if 'last_searched_metrics' not in st.session_state:
            st.session_state.last_searched_metrics = None
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            symbol = st.text_input(
                "Ingresa el símbolo del activo",
                value="",
                help="Ejemplos: AAPL, MSFT, O, T",
                placeholder="AAPL",
                key="search_symbol_input"
            ).upper()
        
        with col2:
            st.write("")  # Espaciado
            st.write("")  # Espaciado
            search_button = st.button("🔍 Buscar", type="primary", use_container_width=True, key="search_button")
        
        # Si se hace clic en buscar, realizar búsqueda
        if symbol and search_button:
            if not ErrorHandler.validate_ticker_symbol(symbol):
                st.error("❌ Símbolo de ticker inválido. Debe ser alfanumérico y tener 1-5 caracteres.")
                st.session_state.last_searched_symbol = None
                st.session_state.last_searched_metrics = None
            else:
                with st.spinner(f"Analizando {symbol}..."):
                    try:
                        metrics = self.analyzer.get_asset_metrics(symbol)
                        
                        if metrics and DataValidator.validate_asset_metrics(metrics):
                            # Guardar en session_state para que persista después del re-render
                            st.session_state.last_searched_symbol = symbol
                            st.session_state.last_searched_metrics = metrics
                            logger.info(f"✅ Búsqueda exitosa para {symbol}, métricas guardadas en session_state")
                        else:
                            st.error(f"❌ No se pudieron obtener datos válidos para {symbol}")
                            st.session_state.last_searched_symbol = None
                            st.session_state.last_searched_metrics = None
                            logger.warning(f"❌ Métricas inválidas para {symbol}")
                            
                    except Exception as e:
                        logger.error(f"Error buscando {symbol}: {e}", exc_info=True)
                        st.error(f"❌ Error al buscar {symbol}: {str(e)}")
                        st.session_state.last_searched_symbol = None
                        st.session_state.last_searched_metrics = None
        
        # Mostrar resultados si hay datos en session_state
        if st.session_state.last_searched_metrics and st.session_state.last_searched_symbol:
            metrics = st.session_state.last_searched_metrics
            symbol = st.session_state.last_searched_symbol
            
            st.markdown("---")
            st.markdown(f"### 📊 Resultados para {symbol}")
            
            # Mostrar métricas principales
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Precio", format_currency(metrics['current_price']))
            with col2:
                st.metric("Dividend Yield", format_percentage(metrics['dividend_yield']))
            with col3:
                st.metric("Dividendo Anual", format_currency(metrics['annual_dividend']))
            with col4:
                freq_emoji = {
                    'mensual': '📅',
                    'trimestral': '📆',
                    'irregular': '⚠️',
                    'sin_dividendos': '❌'
                }
                emoji = freq_emoji.get(metrics['dividend_frequency'], '❓')
                st.metric("Frecuencia", 
                        f"{emoji} {metrics['dividend_frequency'].upper()}")
            
            # Mostrar plataformas si existen
            platforms = self.db.get_platforms(symbol)
            if platforms:
                st.markdown("### 🏪 Plataformas de Compra")
                st.info(f"**Disponible en:** {', '.join(platforms)}")
            else:
                st.markdown("### 🏪 Plataformas de Compra")
                st.info("💡 No hay plataformas asociadas. Ve a '🏪 Gestión de Plataformas' para agregarlas.")
            
            # Detalles adicionales
            with st.expander("📋 Ver detalles completos"):
                st.json(metrics)
            
            # Botón para guardar (siempre visible si hay métricas)
            st.markdown("---")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                save_button = st.button(
                    f"💾 Guardar {symbol} en Base de Datos", 
                    type="primary",
                    use_container_width=True,
                    key="save_to_db_button"
                )
                
                if save_button:
                    logger.info(f"Usuario intenta guardar {symbol} en BD")
                    try:
                        if self.db.upsert_asset(metrics):
                            st.success(f"✅ {symbol} guardado correctamente en la base de datos")
                            logger.info(f"✅ {symbol} guardado exitosamente por el usuario")
                            # Opcional: limpiar session_state después de guardar
                            # st.session_state.last_searched_metrics = None
                        else:
                            st.error("❌ Error al guardar. Revisa los logs para más detalles.")
                            logger.error(f"❌ Fallo al guardar {symbol} por el usuario")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        logger.error(f"❌ Excepción al guardar {symbol}: {e}", exc_info=True)
    
    def view_assets_page(self):
        """Página para ver activos guardados."""
        st.header("📊 Activos Guardados")
        
        # Información de depuración (colapsable)
        with st.expander("🔧 Información de Depuración", expanded=False):
            debug_info = self.db.get_debug_info()
            st.json(debug_info)
            st.write(f"**Ruta de BD:** `{debug_info.get('db_path', 'N/A')}`")
            st.write(f"**Conexión activa:** {debug_info.get('connection_active', False)}")
            st.write(f"**Total de registros:** {debug_info.get('total_records', 0)}")
            if debug_info.get('examples'):
                st.write("**Ejemplos de registros:**")
                st.json(debug_info.get('examples', []))
            
            # Botón de prueba de guardado
            if st.button("🧪 Probar Guardado (AAPL)", key="test_save"):
                st.info("Probando guardado de AAPL...")
                try:
                    test_metrics = self.analyzer.get_asset_metrics("AAPL")
                    if test_metrics:
                        st.json(test_metrics)
                        if DataValidator.validate_asset_metrics(test_metrics):
                            st.success("✅ Métricas válidas")
                            if self.db.upsert_asset(test_metrics):
                                st.success("✅ Guardado exitoso")
                                # Refrescar página para ver el nuevo registro
                                st.rerun()
                            else:
                                st.error("❌ Error al guardar")
                        else:
                            st.error("❌ Métricas inválidas")
                    else:
                        st.error("❌ No se pudieron obtener métricas")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    logger.error(f"Error en prueba de guardado: {e}", exc_info=True)
        
        # Filtros
        col1, col2, col3 = st.columns(3)
        
        with col1:
            filter_freq = st.selectbox(
                "Filtrar por frecuencia:",
                ["Todos", "mensual", "trimestral", "irregular", "sin_dividendos"]
            )
        
        with col2:
            min_yield = st.number_input(
                "Yield mínimo (%)",
                min_value=0.0,
                max_value=100.0,
                value=0.0,
                step=0.1
            )
        
        with col3:
            # Filtro por plataforma
            all_platforms = self.db.get_all_platforms()
            platform_options = ["Todas"] + all_platforms if all_platforms else ["Todas"]
            filter_platform = st.selectbox(
                "Filtrar por plataforma:",
                platform_options
            )
        
        # Obtener activos
        try:
            freq_filter = None if filter_freq == "Todos" else filter_freq
            
            # Filtrar por plataforma si se seleccionó una
            if filter_platform and filter_platform != "Todas":
                assets = self.db.get_assets_by_platform(filter_platform)
                # Aplicar filtro de frecuencia si es necesario
                if freq_filter:
                    assets = [a for a in assets if a.get('dividend_frequency') == freq_filter]
            else:
                assets = self.db.get_all_assets(freq_filter)
            
            logger.info(f"Obtenidos {len(assets)} activos de la BD (filtro: {freq_filter}, plataforma: {filter_platform})")
            
            # Filtrar por yield
            if min_yield > 0:
                assets = [a for a in assets if a.get('dividend_yield', 0) >= min_yield]
                logger.info(f"Después de filtrar por yield >= {min_yield}: {len(assets)} activos")
            
            if assets:
                # Convertir a DataFrame
                df = pd.DataFrame(assets)
                
                # Agregar columna de plataformas formateada
                if 'platforms' in df.columns:
                    df['Plataformas'] = df['platforms'].apply(
                        lambda x: ', '.join([p.strip() for p in str(x).split(',') if p.strip()]) 
                        if x and str(x) != 'nan' else 'Sin plataformas'
                    )
                else:
                    df['Plataformas'] = 'Sin plataformas'
                
                # Seleccionar columnas relevantes
                display_cols = ['symbol', 'name', 'current_price', 'dividend_yield', 
                              'dividend_frequency', 'Plataformas', 'sector', 'last_updated']
                available_cols = [col for col in display_cols if col in df.columns]
                
                # Renombrar columnas para mejor visualización
                df_display = df[available_cols].copy()
                df_display.columns = [col.replace('_', ' ').title() for col in df_display.columns]
                
                st.dataframe(
                    df_display,
                    use_container_width=True
                )
                
                st.success(f"📊 Total: {len(assets)} activos mostrados")
            else:
                st.warning("⚠️ No hay activos que cumplan los filtros. Usa 'Importar Excel' o 'Buscar Activo' para agregar datos.")
                
        except Exception as e:
            logger.error(f"Error obteniendo activos: {e}", exc_info=True)
            st.error(f"❌ Error al obtener activos: {str(e)}")
            st.info("💡 Revisa la información de depuración arriba para más detalles.")
    
    def visualizations_page(self):
        """Página de visualizaciones."""
        st.header("📈 Visualizaciones Financieras")
        
        # Filtros para visualizaciones
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            filter_freq = st.selectbox(
                "Filtrar por frecuencia:",
                ["Todos", "mensual", "trimestral", "irregular"],
                key="viz_freq"
            )
        
        with col2:
            min_yield = st.number_input(
                "Yield mínimo (%)",
                min_value=0.0,
                max_value=100.0,
                value=0.0,
                step=0.1,
                key="viz_yield"
            )
        
        with col3:
            # Filtro por plataforma
            all_platforms = self.db.get_all_platforms()
            platform_options = ["Todas"] + all_platforms if all_platforms else ["Todas"]
            filter_platform = st.selectbox(
                "Filtrar por plataforma:",
                platform_options,
                key="viz_platform"
            )
        
        with col4:
            # Filtro por mes de pago de dividendo
            month_names = {
                0: "Todos los meses",
                1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
                9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
            }
            month_options = [(0, "Todos los meses")] + [(i, month_names[i]) for i in range(1, 13)]
            selected_month = st.selectbox(
                "Filtrar por mes de pago:",
                options=month_options,
                format_func=lambda x: x[1],
                key="viz_payment_month"
            )
            filter_payment_month = selected_month[0] if selected_month[0] > 0 else None
        
        freq_filter = None if filter_freq == "Todos" else filter_freq
        platform_filter = None if filter_platform == "Todas" else filter_platform
        
        # Tabs para diferentes visualizaciones
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "💰 La Búsqueda del Tesoro",
            "📊 Distribución de Yields",
            "🏆 Top Performers",
            "📈 Comparación por Frecuencia",
            "🏪 Activos por Plataforma"
        ])
        
        with tab1:
            st.markdown("""
            ### 💎 La Búsqueda del Tesoro: Guía de Interpretación
            
            Este gráfico te ayuda a encontrar las mejores oportunidades de inversión en dividendos.
            
            #### 📊 **Cómo leer el gráfico:**
            
            **Eje X (Horizontal - Precio):**
            - Muestra el precio actual del activo en USD
            - **Izquierda** = Más barato (mejor para entrar)
            - **Derecha** = Más caro
            
            **Eje Y (Vertical - Yield):**
            - Muestra el porcentaje de retorno por dividendos
            - **Arriba** = Mayor retorno (mejor)
            - **Abajo** = Menor retorno
            
            **Colores y Formas (Leyenda a la derecha):**
            - 🟢 **Círculos verdes**: Pago Mensual (10-12 pagos/año) - Ideal para flujo constante
            - 🔵 **Cuadrados azules**: Pago Trimestral (3-4 pagos/año) - Más común
            - 🟠 **Diamantes naranjas**: Pago Irregular (1-2 pagos/año) - Esporádico
            - 🔴 **X rojas**: Sin dividendos - No paga actualmente
            
            #### 💡 **Zonas del gráfico:**
            
            - **💎 Zona de Gemas (Superior Izquierda)**: Bajo precio + Alto yield = **MEJOR OPCIÓN**
            - **⚠️ Zona a Evitar (Inferior Derecha)**: Alto precio + Bajo yield = **NO EFICIENTE**
            - **Zona Media**: Evaluar caso por caso según tus objetivos
            
            **💡 Tip:** Pasa el mouse sobre cualquier punto para ver detalles del activo.
            """)
            
            try:
                fig = self.visualizer.create_treasure_hunt_scatter(
                    filter_frequency=freq_filter,
                    min_yield=min_yield,
                    filter_platform=platform_filter,
                    filter_payment_month=filter_payment_month
                )
                st.plotly_chart(fig, use_container_width=True, key="treasure_hunt_chart")
            except Exception as e:
                logger.error(f"Error creando gráfico de búsqueda: {e}")
                st.error("❌ Error al crear el gráfico")
        
        with tab2:
            st.markdown("### 📊 Distribución de Dividend Yields")
            try:
                fig = self.visualizer.create_yield_distribution(
                    filter_frequency=freq_filter,
                    filter_platform=platform_filter,
                    filter_payment_month=filter_payment_month
                )
                st.plotly_chart(fig, use_container_width=True, key="yield_distribution_chart")
            except Exception as e:
                logger.error(f"Error creando distribución: {e}")
                st.error("❌ Error al crear el gráfico")
        
        with tab3:
            top_n = st.slider("Número de activos a mostrar", 5, 20, 10, key="top_n_slider")
            st.markdown(f"### 🏆 Top {top_n} Activos por Yield")
            try:
                fig = self.visualizer.create_top_performers(top_n=top_n)
                st.plotly_chart(fig, use_container_width=True, key="top_performers_chart")
            except Exception as e:
                logger.error(f"Error creando top performers: {e}")
                st.error("❌ Error al crear el gráfico")
        
        with tab4:
            st.markdown("### 📈 Comparación por Frecuencia de Dividendos")
            try:
                fig = self.visualizer.create_frequency_comparison(
                    filter_platform=platform_filter,
                    filter_payment_month=filter_payment_month
                )
                st.plotly_chart(fig, use_container_width=True, key="frequency_comparison_chart")
            except Exception as e:
                logger.error(f"Error creando comparación: {e}")
                st.error("❌ Error al crear el gráfico")
        
        with tab5:
            st.markdown("### 🏪 Distribución de Activos por Plataforma")
            st.info("Este gráfico muestra cuántos activos están disponibles en cada plataforma.")
            try:
                fig = self.visualizer.create_platform_distribution()
                st.plotly_chart(fig, use_container_width=True, key="platform_distribution_chart")
            except Exception as e:
                logger.error(f"Error creando distribución por plataforma: {e}")
                st.error("❌ Error al crear el gráfico")
            except Exception as e:
                logger.error(f"Error creando comparación: {e}")
                st.error("❌ Error al crear el gráfico")
    
    def stats_page(self):
        """Página de estadísticas."""
        st.header("ℹ️ Estadísticas del Portfolio")
        
        try:
            stats = self.db.get_stats()
            
            if stats:
                # Métricas principales
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Activos", stats.get('total_assets', 0))
                
                with col2:
                    st.metric("Yield Promedio", 
                             format_percentage(stats.get('average_yield', 0)))
                
                with col3:
                    freq_dist = stats.get('frequency_distribution', {})
                    st.metric("Frecuencias Diferentes", len(freq_dist))
                
                # Distribución por frecuencia
                if freq_dist:
                    st.markdown("### Distribución por Frecuencia")
                    freq_df = pd.DataFrame(
                        list(freq_dist.items()),
                        columns=['Frecuencia', 'Cantidad']
                    )
                    st.bar_chart(freq_df.set_index('Frecuencia'))
                    
                    # Tabla detallada
                    with st.expander("Ver tabla detallada"):
                        st.dataframe(freq_df, use_container_width=True)
            else:
                st.warning("⚠️ No hay estadísticas disponibles. Importa algunos activos primero.")
                
        except Exception as e:
            logger.error(f"Error obteniendo estadísticas: {e}")
            st.error("❌ Error al obtener estadísticas")
    
    def maintenance_page(self):
        """Página de mantenimiento con funciones de actualización y búsqueda automática."""
        st.header("🔧 Mantenimiento de Base de Datos")
        
        st.markdown("""
        Esta sección te permite mantener actualizada tu base de datos:
        - **Actualizar Activos**: Actualiza la información de todos los activos ya registrados
        - **Búsqueda Automática**: Busca y agrega nuevos activos automáticamente
        """)
        
        # Tabs para las dos funciones
        tab1, tab2 = st.tabs(["🔄 Actualizar Activos Existentes", "🔍 Búsqueda Automática de Nuevos Activos"])
        
        with tab1:
            self._update_all_assets_tab()
        
        with tab2:
            self._auto_search_new_assets_tab()
    
    def _update_all_assets_tab(self):
        """Tab para actualizar todos los activos existentes."""
        st.markdown("### 🔄 Actualizar Información de Activos Existentes")
        
        st.info("""
        Esta función actualizará la información de **todos** los activos que ya están 
        almacenados en la base de datos. Esto incluye precios, dividendos, yields y 
        frecuencias de pago.
        """)
        
        # Obtener todos los símbolos
        all_symbols = self.db.get_all_symbols()
        
        if not all_symbols:
            st.warning("⚠️ No hay activos en la base de datos para actualizar.")
            st.info("💡 Usa 'Importar Excel' o 'Buscar Activo' para agregar activos primero.")
            return
        
        st.write(f"**Total de activos a actualizar:** {len(all_symbols)}")
        
        # Mostrar lista de símbolos
        with st.expander("Ver lista de activos a actualizar"):
            st.write(", ".join(all_symbols))
        
        # Botón para actualizar
        if st.button("🚀 Actualizar Todos los Activos", type="primary", key="update_all_button"):
            self._process_update_all_assets(all_symbols)
    
    def _process_update_all_assets(self, symbols: List[str]):
        """Procesa la actualización de todos los activos."""
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        success_count = 0
        error_count = 0
        results = []
        
        for i, symbol in enumerate(symbols):
            status_text.text(f"Actualizando {symbol}... ({i+1}/{len(symbols)})")
            
            try:
                # Obtener nuevas métricas
                metrics = self.analyzer.get_asset_metrics(symbol)
                
                if metrics and DataValidator.validate_asset_metrics(metrics):
                    # Actualizar en BD (upsert actualizará el registro existente)
                    if self.db.upsert_asset(metrics):
                        success_count += 1
                        results.append({
                            'Símbolo': symbol,
                            'Estado': '✅ Actualizado',
                            'Precio': format_currency(metrics.get('current_price', 0)),
                            'Yield': format_percentage(metrics.get('dividend_yield', 0)),
                            'Frecuencia': metrics.get('dividend_frequency', 'N/A')
                        })
                        logger.info(f"✅ Actualizado: {symbol}")
                    else:
                        error_count += 1
                        results.append({
                            'Símbolo': symbol,
                            'Estado': '❌ Error BD',
                            'Precio': 'N/A',
                            'Yield': 'N/A',
                            'Frecuencia': 'N/A'
                        })
                else:
                    error_count += 1
                    results.append({
                        'Símbolo': symbol,
                        'Estado': '❌ Sin datos',
                        'Precio': 'N/A',
                        'Yield': 'N/A',
                        'Frecuencia': 'N/A'
                    })
                    logger.warning(f"❌ No se pudieron obtener datos para {symbol}")
                    
            except Exception as e:
                logger.error(f"Error actualizando {symbol}: {e}", exc_info=True)
                error_count += 1
                results.append({
                    'Símbolo': symbol,
                    'Estado': f'❌ Error: {str(e)[:30]}',
                    'Precio': 'N/A',
                    'Yield': 'N/A',
                    'Frecuencia': 'N/A'
                })
            
            progress_bar.progress((i + 1) / len(symbols))
        
        status_text.empty()
        progress_bar.empty()
        
        # Mostrar resumen
        st.success(f"✅ Actualización completada: {success_count} exitosos, {error_count} con errores")
        
        # Mostrar tabla de resultados
        if results:
            results_df = pd.DataFrame(results)
            st.dataframe(results_df, use_container_width=True)
    
    def _auto_search_new_assets_tab(self):
        """Tab para búsqueda automática de nuevos activos."""
        st.markdown("### 🔍 Búsqueda Automática de Nuevos Activos")
        
        st.info("""
        Esta función busca automáticamente nuevos activos y los agrega a la base de datos
        si no están ya almacenados. Puedes usar una lista predefinida o ingresar tus propios tickers.
        """)
        
        # Lista predefinida de tickers populares
        popular_tickers = [
            # Tech
            'AAPL', 'MSFT', 'GOOGL', 'AMZN', 'META', 'NVDA', 'TSLA', 'NFLX',
            # Dividendos
            'JNJ', 'PG', 'KO', 'PEP', 'WMT', 'VZ', 'T', 'XOM', 'CVX',
            # REITs (Mensuales)
            'O', 'STAG', 'AGNC', 'MAIN', 'SPG', 'AMT',
            # ETFs
            'SPY', 'VOO', 'SCHD', 'VYM', 'DIV', 'HDV'
        ]
        
        # Opciones de búsqueda
        search_mode = st.radio(
            "Modo de búsqueda:",
            ["📋 Lista Predefinida (Tickers Populares)", "✏️ Lista Personalizada"],
            key="search_mode"
        )
        
        if search_mode == "📋 Lista Predefinida (Tickers Populares)":
            st.write(f"**Tickers a buscar:** {len(popular_tickers)} activos populares")
            with st.expander("Ver lista predefinida"):
                st.write(", ".join(popular_tickers))
            
            tickers_to_search = popular_tickers
        
        else:  # Lista personalizada
            st.markdown("""
            **Ingresa los símbolos de tickers separados por comas o uno por línea:**
            """)
            tickers_input = st.text_area(
                "Tickers a buscar",
                placeholder="AAPL, MSFT, GOOGL\nO uno por línea:\nAAPL\nMSFT\nGOOGL",
                height=150,
                key="custom_tickers_input"
            )
            
            if tickers_input:
                # Procesar input: separar por comas o líneas
                tickers_list = []
                for line in tickers_input.split('\n'):
                    for ticker in line.split(','):
                        ticker = ticker.strip().upper()
                        if ticker and ErrorHandler.validate_ticker_symbol(ticker):
                            tickers_list.append(ticker)
                
                tickers_to_search = list(set(tickers_list))  # Eliminar duplicados
                st.write(f"**Tickers válidos encontrados:** {len(tickers_to_search)}")
                if tickers_to_search:
                    st.write(", ".join(tickers_to_search))
            else:
                tickers_to_search = []
                st.warning("⚠️ Ingresa al menos un ticker para buscar")
        
        # Obtener símbolos ya almacenados
        existing_symbols = set(self.db.get_all_symbols())
        
        if tickers_to_search:
            # Filtrar solo los que no están en la BD
            new_tickers = [t for t in tickers_to_search if t not in existing_symbols]
            already_stored = [t for t in tickers_to_search if t in existing_symbols]
            
            if already_stored:
                st.info(f"ℹ️ {len(already_stored)} tickers ya están almacenados: {', '.join(already_stored)}")
            
            if new_tickers:
                st.success(f"✅ {len(new_tickers)} nuevos tickers para buscar: {', '.join(new_tickers)}")
                
                # Botón para buscar
                if st.button("🚀 Buscar y Agregar Nuevos Activos", type="primary", key="auto_search_button"):
                    self._process_auto_search(new_tickers)
            else:
                st.warning("⚠️ Todos los tickers ya están almacenados en la base de datos.")
    
    def _process_auto_search(self, symbols: List[str]):
        """Procesa la búsqueda automática de nuevos activos."""
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        success_count = 0
        error_count = 0
        results = []
        
        for i, symbol in enumerate(symbols):
            status_text.text(f"Buscando {symbol}... ({i+1}/{len(symbols)})")
            
            try:
                metrics = self.analyzer.get_asset_metrics(symbol)
                
                if metrics and DataValidator.validate_asset_metrics(metrics):
                    # Verificar que realmente no existe (doble verificación)
                    existing = self.db.get_asset(symbol)
                    if existing:
                        results.append({
                            'Símbolo': symbol,
                            'Estado': '⚠️ Ya existe',
                            'Nombre': metrics.get('name', 'N/A'),
                            'Yield': format_percentage(metrics.get('dividend_yield', 0)),
                            'Frecuencia': metrics.get('dividend_frequency', 'N/A')
                        })
                    else:
                        # Guardar nuevo activo
                        if self.db.upsert_asset(metrics):
                            success_count += 1
                            results.append({
                                'Símbolo': symbol,
                                'Estado': '✅ Agregado',
                                'Nombre': metrics.get('name', 'N/A'),
                                'Yield': format_percentage(metrics.get('dividend_yield', 0)),
                                'Frecuencia': metrics.get('dividend_frequency', 'N/A')
                            })
                            logger.info(f"✅ Nuevo activo agregado: {symbol}")
                        else:
                            error_count += 1
                            results.append({
                                'Símbolo': symbol,
                                'Estado': '❌ Error BD',
                                'Nombre': 'N/A',
                                'Yield': 'N/A',
                                'Frecuencia': 'N/A'
                            })
                else:
                    error_count += 1
                    results.append({
                        'Símbolo': symbol,
                        'Estado': '❌ Sin datos',
                        'Nombre': 'N/A',
                        'Yield': 'N/A',
                        'Frecuencia': 'N/A'
                    })
                    
            except Exception as e:
                logger.error(f"Error buscando {symbol}: {e}", exc_info=True)
                error_count += 1
                results.append({
                    'Símbolo': symbol,
                    'Estado': f'❌ Error: {str(e)[:30]}',
                    'Nombre': 'N/A',
                    'Yield': 'N/A',
                    'Frecuencia': 'N/A'
                })
            
            progress_bar.progress((i + 1) / len(symbols))
        
        status_text.empty()
        progress_bar.empty()
        
        # Mostrar resumen
        st.success(f"✅ Búsqueda completada: {success_count} nuevos activos agregados, {error_count} con errores")
        
        # Mostrar tabla de resultados
        if results:
            results_df = pd.DataFrame(results)
            st.dataframe(results_df, use_container_width=True)
            
            if success_count > 0:
                st.info("💡 Ve a '📊 Ver Activos' para ver los nuevos activos agregados.")
    
    def platforms_management_page(self):
        """Página para gestionar plataformas de compra de activos."""
        st.header("🏪 Gestión de Plataformas de Compra")
        
        st.markdown("""
        Asocia acciones con las plataformas donde se pueden comprar. Esto te ayudará a saber
        dónde adquirir cada activo.
        
        **Plataformas comunes:** PREX, REVOLUT, IBKR, Interactive Brokers, eToro, etc.
        """)
        
        # Tabs para diferentes métodos de asociación
        tab1, tab2, tab3 = st.tabs([
            "➕ Asociación Individual", 
            "📋 Importar desde Lista", 
            "📊 Importar desde Excel"
        ])
        
        with tab1:
            self._individual_platform_association()
        
        with tab2:
            self._import_platforms_from_list()
        
        with tab3:
            self._import_platforms_from_excel()
    
    def _individual_platform_association(self):
        """Tab para asociar plataformas de forma individual con edición mejorada."""
        st.markdown("### ➕ Editar Plataformas de un Activo")
        
        # Obtener lista de activos
        all_assets = self.db.get_all_assets()
        
        if not all_assets:
            st.warning("⚠️ No hay activos en la base de datos.")
            st.info("💡 Agrega activos primero usando 'Buscar Activo' o 'Importar Excel'")
            return
        
        # Selector de activo
        asset_options = {f"{a['symbol']} - {a.get('name', 'N/A')}": a['symbol'] 
                        for a in all_assets}
        
        selected_asset_label = st.selectbox(
            "Selecciona un activo:",
            options=list(asset_options.keys()),
            key="platform_asset_selector"
        )
        
        selected_symbol = asset_options[selected_asset_label]
        
        # Obtener plataformas actuales
        current_platforms = set(self.db.get_platforms(selected_symbol))
        
        # Obtener todas las plataformas disponibles en la BD
        all_available_platforms = self.db.get_all_platforms()
        
        st.write(f"**Activo seleccionado:** {selected_symbol}")
        
        # Mostrar plataformas actuales
        if current_platforms:
            st.success(f"**Plataformas actuales:** {', '.join(sorted(current_platforms))}")
        else:
            st.info("**Plataformas actuales:** Ninguna")
        
        st.markdown("---")
        st.markdown("### ✏️ Editar Plataformas")
        
        # Opción 1: Checkboxes con plataformas existentes
        if all_available_platforms:
            st.markdown("**Selecciona de plataformas existentes:**")
            selected_existing = st.multiselect(
                "Plataformas disponibles:",
                options=all_available_platforms,
                default=list(current_platforms.intersection(set(all_available_platforms))),
                key="platforms_multiselect"
            )
        else:
            selected_existing = []
            st.info("💡 No hay plataformas en la BD aún. Agrega nuevas plataformas abajo.")
        
        # Opción 2: Agregar nuevas plataformas
        st.markdown("**O agrega nuevas plataformas:**")
        new_platforms_input = st.text_input(
            "Nuevas plataformas (separadas por comas)",
            key="new_platforms_input",
            help="Ej: PREX, REVOLUT, IBKR"
        )
        
        # Combinar selecciones
        if new_platforms_input.strip():
            new_platforms = [p.strip().upper() for p in new_platforms_input.split(',') if p.strip()]
            final_platforms = list(set(selected_existing + new_platforms))
        else:
            final_platforms = list(selected_existing)
        
        # Mostrar preview
        if final_platforms:
            st.info(f"**Plataformas a guardar:** {', '.join(sorted(final_platforms))}")
        else:
            st.warning("⚠️ No hay plataformas seleccionadas. Se eliminarán todas las plataformas del activo.")
        
        # Botones de acción
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("💾 Guardar Cambios", type="primary", key="save_platforms_individual", use_container_width=True):
                if self.db.update_platforms(selected_symbol, final_platforms):
                    st.success(f"✅ Plataformas actualizadas para {selected_symbol}")
                    logger.info(f"Plataformas actualizadas para {selected_symbol}: {final_platforms}")
                    st.rerun()
                else:
                    st.error(f"❌ Error al actualizar plataformas para {selected_symbol}")
        
        with col2:
            if st.button("🗑️ Eliminar Todas", key="clear_platforms_individual", use_container_width=True):
                if self.db.update_platforms(selected_symbol, []):
                    st.success(f"✅ Todas las plataformas eliminadas para {selected_symbol}")
                    st.rerun()
                else:
                    st.error(f"❌ Error al eliminar plataformas")
    
    def _import_platforms_from_list(self):
        """Tab para importar plataformas desde una lista de texto."""
        st.markdown("### 📋 Importar Plataformas desde Lista de Texto")
        
        st.markdown("""
        **Formato esperado:**
        ```
        SYMBOL: PLATFORM1, PLATFORM2, PLATFORM3
        AAPL: PREX, REVOLUT
        MSFT: PREX, IBKR
        O: PREX
        ```
        
        O un símbolo por línea con plataformas separadas por comas.
        """)
        
        platforms_text = st.text_area(
            "Pega tu lista aquí:",
            height=200,
            placeholder="AAPL: PREX, REVOLUT\nMSFT: PREX, IBKR\nO: PREX",
            key="platforms_list_input"
        )
        
        if st.button("📥 Procesar Lista", type="primary", key="process_platforms_list"):
            if platforms_text.strip():
                self._process_platforms_list(platforms_text)
            else:
                st.warning("⚠️ Ingresa una lista de plataformas")
    
    def _process_platforms_list(self, text: str):
        """Procesa una lista de texto con asociaciones símbolo:plataformas."""
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        results = []
        success_count = 0
        error_count = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, line in enumerate(lines):
            status_text.text(f"Procesando línea {i+1}/{len(lines)}...")
            
            try:
                # Formato: SYMBOL: PLATFORM1, PLATFORM2
                if ':' in line:
                    parts = line.split(':', 1)
                    symbol = parts[0].strip().upper()
                    platforms_str = parts[1].strip()
                else:
                    # Formato alternativo: SYMBOL PLATFORM1, PLATFORM2
                    parts = line.split(None, 1)
                    if len(parts) == 2:
                        symbol = parts[0].strip().upper()
                        platforms_str = parts[1].strip()
                    else:
                        error_count += 1
                        results.append({
                            'Línea': line[:50],
                            'Estado': '❌ Formato inválido',
                            'Plataformas': 'N/A'
                        })
                        continue
                
                # Validar símbolo
                if not ErrorHandler.validate_ticker_symbol(symbol):
                    error_count += 1
                    results.append({
                        'Símbolo': symbol,
                        'Estado': '❌ Símbolo inválido',
                        'Plataformas': platforms_str
                    })
                    continue
                
                # Verificar que el activo existe
                asset = self.db.get_asset(symbol)
                if not asset:
                    error_count += 1
                    results.append({
                        'Símbolo': symbol,
                        'Estado': '❌ Activo no encontrado',
                        'Plataformas': platforms_str
                    })
                    continue
                
                # Procesar plataformas
                platforms_list = [p.strip().upper() for p in platforms_str.split(',') if p.strip()]
                
                if platforms_list:
                    if self.db.update_platforms(symbol, platforms_list):
                        success_count += 1
                        results.append({
                            'Símbolo': symbol,
                            'Estado': '✅ Actualizado',
                            'Plataformas': ', '.join(platforms_list)
                        })
                    else:
                        error_count += 1
                        results.append({
                            'Símbolo': symbol,
                            'Estado': '❌ Error BD',
                            'Plataformas': platforms_str
                        })
                else:
                    error_count += 1
                    results.append({
                        'Símbolo': symbol,
                        'Estado': '❌ Sin plataformas',
                        'Plataformas': platforms_str
                    })
                    
            except Exception as e:
                logger.error(f"Error procesando línea: {line}, Error: {e}")
                error_count += 1
                results.append({
                    'Línea': line[:50],
                    'Estado': f'❌ Error: {str(e)[:30]}',
                    'Plataformas': 'N/A'
                })
            
            progress_bar.progress((i + 1) / len(lines))
        
        status_text.empty()
        progress_bar.empty()
        
        # Mostrar resultados
        st.success(f"✅ Procesamiento completado: {success_count} exitosos, {error_count} con errores")
        
        if results:
            results_df = pd.DataFrame(results)
            st.dataframe(results_df, use_container_width=True)
    
    def _import_platforms_from_excel(self):
        """Tab para importar plataformas desde Excel."""
        st.markdown("### 📊 Importar Plataformas desde Excel")
        
        st.markdown("""
        **Formato del archivo Excel:**
        - **Columna A:** Símbolo del activo (ej: AAPL, MSFT)
        - **Columna B:** Plataformas separadas por comas (ej: PREX, REVOLUT)
        
        El archivo debe tener encabezados en la primera fila.
        """)
        
        uploaded_file = st.file_uploader(
            "Selecciona archivo Excel (.xlsx)",
            type=['xlsx'],
            key="platforms_excel_uploader",
            help="Columna A: Símbolo, Columna B: Plataformas (separadas por comas)"
        )
        
        if uploaded_file is not None:
            try:
                # Leer Excel
                df = pd.read_excel(uploaded_file, engine='openpyxl')
                
                st.success(f"✅ Archivo cargado: {len(df)} filas")
                
                # Mostrar preview
                with st.expander("Vista previa del archivo"):
                    st.dataframe(df.head(10))
                
                # Verificar columnas
                if len(df.columns) < 2:
                    st.error("❌ El archivo debe tener al menos 2 columnas (Símbolo y Plataformas)")
                    return
                
                # Procesar
                if st.button("📥 Procesar Excel", type="primary", key="process_platforms_excel"):
                    self._process_platforms_excel(df)
                    
            except Exception as e:
                st.error(f"❌ Error leyendo archivo: {str(e)}")
                logger.error(f"Error leyendo Excel de plataformas: {e}", exc_info=True)
    
    def _process_platforms_excel(self, df: pd.DataFrame):
        """Procesa un DataFrame de Excel con asociaciones símbolo:plataformas."""
        results = []
        success_count = 0
        error_count = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, row in df.iterrows():
            status_text.text(f"Procesando fila {i+1}/{len(df)}...")
            
            try:
                # Primera columna: símbolo, Segunda columna: plataformas
                symbol = str(row.iloc[0]).strip().upper()
                platforms_str = str(row.iloc[1]) if len(row) > 1 else ""
                
                # Validar símbolo
                if not symbol or symbol == 'nan':
                    error_count += 1
                    results.append({
                        'Fila': i+2,
                        'Símbolo': 'N/A',
                        'Estado': '❌ Símbolo vacío',
                        'Plataformas': 'N/A'
                    })
                    continue
                
                if not ErrorHandler.validate_ticker_symbol(symbol):
                    error_count += 1
                    results.append({
                        'Fila': i+2,
                        'Símbolo': symbol,
                        'Estado': '❌ Símbolo inválido',
                        'Plataformas': platforms_str
                    })
                    continue
                
                # Verificar que el activo existe
                asset = self.db.get_asset(symbol)
                if not asset:
                    error_count += 1
                    results.append({
                        'Fila': i+2,
                        'Símbolo': symbol,
                        'Estado': '❌ Activo no encontrado',
                        'Plataformas': platforms_str
                    })
                    continue
                
                # Procesar plataformas
                if platforms_str and platforms_str != 'nan':
                    platforms_list = [p.strip().upper() for p in str(platforms_str).split(',') if p.strip()]
                else:
                    platforms_list = []
                
                if platforms_list:
                    if self.db.update_platforms(symbol, platforms_list):
                        success_count += 1
                        results.append({
                            'Fila': i+2,
                            'Símbolo': symbol,
                            'Estado': '✅ Actualizado',
                            'Plataformas': ', '.join(platforms_list)
                        })
                    else:
                        error_count += 1
                        results.append({
                            'Fila': i+2,
                            'Símbolo': symbol,
                            'Estado': '❌ Error BD',
                            'Plataformas': platforms_str
                        })
                else:
                    error_count += 1
                    results.append({
                        'Fila': i+2,
                        'Símbolo': symbol,
                        'Estado': '❌ Sin plataformas',
                        'Plataformas': platforms_str
                    })
                    
            except Exception as e:
                logger.error(f"Error procesando fila {i+1}: {e}")
                error_count += 1
                results.append({
                    'Fila': i+2,
                    'Símbolo': 'N/A',
                    'Estado': f'❌ Error: {str(e)[:30]}',
                    'Plataformas': 'N/A'
                })
            
            progress_bar.progress((i + 1) / len(df))
        
        status_text.empty()
        progress_bar.empty()
        
        # Mostrar resultados
        st.success(f"✅ Procesamiento completado: {success_count} exitosos, {error_count} con errores")
        
        if results:
            results_df = pd.DataFrame(results)
            st.dataframe(results_df, use_container_width=True)
    
    def portfolio_builder_page(self):
        """Página para construir y visualizar un portfolio con calendario de dividendos."""
        st.header("💼 Constructor de Portfolio")
        
        st.markdown("""
        Construye tu portfolio seleccionando acciones y visualiza un calendario anual
        de dividendos con las ganancias proyectadas por mes.
        """)
        
        # Inicializar session_state para el portfolio
        if 'portfolio_selected' not in st.session_state:
            st.session_state.portfolio_selected = []
        if 'portfolio_shares' not in st.session_state:
            st.session_state.portfolio_shares = {}  # {symbol: cantidad}
        if 'portfolio_tax_rates' not in st.session_state:
            st.session_state.portfolio_tax_rates = {}  # {symbol: porcentaje_impuesto}
        if 'portfolio_budget' not in st.session_state:
            st.session_state.portfolio_budget = 0.0
        if 'portfolio_loaded_name' not in st.session_state:
            st.session_state.portfolio_loaded_name = None  # Nombre del portfolio cargado
        
        # Tabs principales
        tab1, tab2, tab3, tab4 = st.tabs([
            "📋 Seleccionar Acciones",
            "📅 Calendario de Dividendos",
            "💰 Resumen Financiero",
            "💾 Portfolios Guardados"
        ])
        
        with tab1:
            self._portfolio_selection_tab()
        
        with tab2:
            self._portfolio_calendar_tab()
        
        with tab3:
            self._portfolio_summary_tab()
        
        with tab4:
            self._portfolio_saved_tab()
    
    def _portfolio_selection_tab(self):
        """Tab para seleccionar acciones del portfolio."""
        st.markdown("### 📋 Seleccionar Acciones para tu Portfolio")
        
        # Obtener todas las acciones
        all_assets = self.db.get_all_assets()
        
        if not all_assets:
            st.warning("⚠️ No hay activos en la base de datos.")
            st.info("💡 Agrega activos primero usando 'Buscar Activo' o 'Importar Excel'")
            return
        
        # Filtros para selección
        col1, col2, col3 = st.columns(3)
        
        with col1:
            all_platforms = self.db.get_all_platforms()
            platform_options = ["Todas", "Sin plataforma"] + (all_platforms if all_platforms else [])
            filter_platform = st.selectbox(
                "Filtrar por plataforma:",
                platform_options,
                key="portfolio_platform_filter"
            )
        
        with col2:
            filter_freq = st.selectbox(
                "Filtrar por frecuencia:",
                ["Todas", "mensual", "trimestral", "irregular"],
                key="portfolio_freq_filter"
            )
        
        with col3:
            min_yield = st.number_input(
                "Yield mínimo (%)",
                min_value=0.0,
                max_value=100.0,
                value=0.0,
                step=0.1,
                key="portfolio_min_yield"
            )
        
        # Filtrar activos
        filtered_assets = all_assets.copy()
        
        if filter_platform and filter_platform != "Todas":
            if filter_platform == "Sin plataforma":
                filtered_assets = [a for a in filtered_assets 
                                 if not a.get('platforms') or str(a.get('platforms', '')) == 'nan']
            else:
                filtered_assets = self.db.get_assets_by_platform(filter_platform)
                if filter_freq != "Todas":
                    filtered_assets = [a for a in filtered_assets 
                                     if a.get('dividend_frequency') == filter_freq]
        
        if filter_freq != "Todas" and filter_platform == "Todas":
            filtered_assets = [a for a in filtered_assets 
                             if a.get('dividend_frequency') == filter_freq]
        
        if min_yield > 0:
            filtered_assets = [a for a in filtered_assets 
                             if a.get('dividend_yield', 0) >= min_yield]
        
        st.write(f"**Activos disponibles:** {len(filtered_assets)}")
        
        # Tabla de selección
        if filtered_assets:
            # Crear lista de opciones para multiselect
            asset_options = {}
            for asset in filtered_assets:
                symbol = asset['symbol']
                name = asset.get('name', 'N/A')
                yield_val = asset.get('dividend_yield', 0)
                freq = asset.get('dividend_frequency', 'N/A')
                platforms = asset.get('platforms', '')
                if platforms and str(platforms) != 'nan':
                    platforms_str = ', '.join([p.strip() for p in str(platforms).split(',') if p.strip()])
                else:
                    platforms_str = 'Sin plataforma'
                
                label = f"{symbol} - {name} | Yield: {yield_val:.2f}% | {freq} | {platforms_str}"
                asset_options[label] = symbol
            
            # Mostrar información si hay un portfolio cargado
            if st.session_state.get('portfolio_loaded_name'):
                st.info(f"📂 Portfolio cargado: **{st.session_state.portfolio_loaded_name}** ({len(st.session_state.portfolio_selected)} acciones)")
            
            # Multi-select - Asegurar que los valores por defecto coincidan con session_state
            default_labels = []
            if st.session_state.portfolio_selected:
                # Crear mapeo inverso: symbol -> label (normalizar a mayúsculas para comparación)
                symbol_to_label = {}
                for label, symbol in asset_options.items():
                    symbol_upper = symbol.upper()
                    symbol_to_label[symbol_upper] = label
                    symbol_to_label[symbol] = label  # También mapear sin normalizar
                
                # Buscar labels para los símbolos seleccionados (normalizar a mayúsculas)
                for symbol in st.session_state.portfolio_selected:
                    symbol_upper = symbol.upper() if isinstance(symbol, str) else str(symbol).upper()
                    if symbol_upper in symbol_to_label:
                        default_labels.append(symbol_to_label[symbol_upper])
                    elif symbol in symbol_to_label:
                        default_labels.append(symbol_to_label[symbol])
            
            # Usar una clave única que incluya el nombre del portfolio cargado para forzar actualización
            multiselect_key = f"portfolio_multiselect_{st.session_state.get('portfolio_loaded_name', 'default')}"
            
            selected_labels = st.multiselect(
                "Selecciona acciones para tu portfolio:",
                options=list(asset_options.keys()),
                default=default_labels,
                key=multiselect_key
            )
            
            # Actualizar session_state con las selecciones actuales (normalizar a mayúsculas)
            current_selected = [asset_options[label] for label in selected_labels]
            current_selected = [s.upper() if isinstance(s, str) else str(s).upper() for s in current_selected]
            
            # Solo actualizar si realmente hay cambios o si hay un portfolio cargado
            if st.session_state.get('portfolio_loaded_name') or set(current_selected) != set(st.session_state.portfolio_selected):
                st.session_state.portfolio_selected = current_selected
            
            # Mostrar selección actual y permitir asignar cantidades
            if st.session_state.portfolio_selected:
                st.success(f"✅ {len(st.session_state.portfolio_selected)} acciones seleccionadas")
                
                # Debug: mostrar qué acciones están seleccionadas
                if st.session_state.get('portfolio_loaded_name'):
                    st.caption(f"📋 Acciones del portfolio: {', '.join(st.session_state.portfolio_selected[:10])}")
                    if len(st.session_state.portfolio_selected) > 10:
                        st.caption(f"... y {len(st.session_state.portfolio_selected) - 10} más")
                
                st.markdown("---")
                st.markdown("### 📊 Asignar Cantidad de Acciones")
                st.info("💡 Especifica cuántas acciones de cada tipo quieres en tu portfolio")
                
                # Obtener datos de las acciones seleccionadas
                # Normalizar símbolos a mayúsculas para comparación
                selected_symbols_upper = [s.upper() if isinstance(s, str) else str(s).upper() for s in st.session_state.portfolio_selected]
                selected_assets = [a for a in filtered_assets 
                                 if a['symbol'].upper() in selected_symbols_upper]
                
                # Verificar que todas las acciones seleccionadas estén disponibles
                available_symbols = [a['symbol'].upper() for a in filtered_assets]
                missing_symbols = [s for s in selected_symbols_upper if s not in available_symbols]
                if missing_symbols:
                    st.warning(f"⚠️ Algunas acciones del portfolio no están disponibles en los filtros actuales: {', '.join(missing_symbols)}")
                    st.info("💡 Ajusta los filtros (plataforma, frecuencia, etc.) para ver todas las acciones del portfolio")
                
                if not selected_assets:
                    st.error("❌ No se encontraron acciones disponibles. Ajusta los filtros para ver las acciones del portfolio.")
                    return
                
                # Crear tabla editable para asignar cantidades e impuestos
                portfolio_data = []
                for asset in selected_assets:
                    symbol = asset['symbol']
                    # Buscar en session_state usando el símbolo en mayúsculas (normalizado)
                    symbol_key = symbol.upper()
                    current_shares = (st.session_state.portfolio_shares.get(symbol_key, 
                                                                          st.session_state.portfolio_shares.get(symbol, 0)))
                    current_tax_rate = (st.session_state.portfolio_tax_rates.get(symbol_key,
                                                                                st.session_state.portfolio_tax_rates.get(symbol, 0.0)))
                    
                    col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 2])
                    
                    with col1:
                        st.write(f"**{symbol}**")
                        st.caption(asset.get('name', 'N/A'))
                    
                    with col2:
                        st.write(f"Precio: {format_currency(asset.get('current_price', 0))}")
                    
                    with col3:
                        shares = st.number_input(
                            "Cantidad:",
                            min_value=0,
                            value=int(current_shares),
                            step=1,
                            key=f"shares_{symbol}"
                        )
                        st.session_state.portfolio_shares[symbol] = shares
                    
                    with col4:
                        tax_rate = st.number_input(
                            "Impuesto (%):",
                            min_value=0.0,
                            max_value=100.0,
                            value=float(current_tax_rate),
                            step=0.1,
                            format="%.1f",
                            key=f"tax_{symbol}",
                            help="Porcentaje de descuento de impuesto sobre dividendos"
                        )
                        st.session_state.portfolio_tax_rates[symbol] = tax_rate
                    
                    with col5:
                        cost = asset.get('current_price', 0) * shares
                        st.metric("Costo", format_currency(cost))
                    
                    # Calcular dividendos después de impuestos
                    annual_dividend_before_tax = (asset.get('annual_dividend', 0) or 0) * shares
                    tax_amount = annual_dividend_before_tax * (tax_rate / 100)
                    annual_dividend_after_tax = annual_dividend_before_tax - tax_amount
                    
                    portfolio_data.append({
                        'symbol': symbol,
                        'name': asset.get('name', 'N/A'),
                        'price': asset.get('current_price', 0),
                        'shares': shares,
                        'total_cost': cost,
                        'annual_dividend': annual_dividend_before_tax,
                        'annual_dividend_after_tax': annual_dividend_after_tax,
                        'tax_rate': tax_rate,
                        'tax_amount': tax_amount,
                        'yield': asset.get('dividend_yield', 0)
                    })
                    
                    st.markdown("---")
                
                # Resumen del portfolio
                if portfolio_data:
                    total_portfolio_cost = sum(item['total_cost'] for item in portfolio_data)
                    total_annual_dividends_before_tax = sum(item['annual_dividend'] for item in portfolio_data)
                    total_annual_dividends_after_tax = sum(item['annual_dividend_after_tax'] for item in portfolio_data)
                    total_tax_amount = sum(item['tax_amount'] for item in portfolio_data)
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("💰 Costo Total del Portfolio", format_currency(total_portfolio_cost))
                    
                    with col2:
                        st.metric("💵 Dividendos Anuales (bruto)", format_currency(total_annual_dividends_before_tax))
                    
                    with col3:
                        st.metric("💵 Dividendos Anuales (neto)", format_currency(total_annual_dividends_after_tax))
                    
                    with col4:
                        st.metric("📊 Impuestos Totales", format_currency(total_tax_amount))
                    
                    # Yield después de impuestos
                    if total_portfolio_cost > 0:
                        portfolio_yield_after_tax = (total_annual_dividends_after_tax / total_portfolio_cost) * 100
                        st.metric("📈 Yield del Portfolio (neto)", format_percentage(portfolio_yield_after_tax))
                    
                    # Botones para guardar portfolio
                    st.markdown("---")
                    st.markdown("### 💾 Guardar Portfolio")
                    
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        portfolio_name = st.text_input(
                            "Nombre del portfolio:",
                            key="portfolio_name_input",
                            placeholder="Ej: Portfolio Conservador 2024"
                        )
                        portfolio_description = st.text_area(
                            "Descripción (opcional):",
                            key="portfolio_description_input",
                            placeholder="Descripción del portfolio...",
                            height=80
                        )
                    
                    with col2:
                        st.write("")  # Espaciado
                        st.write("")  # Espaciado
                        if st.button("💾 Guardar Portfolio", key="save_portfolio_btn", use_container_width=True):
                            if portfolio_name:
                                success = self.db.save_portfolio(
                                    name=portfolio_name,
                                    description=portfolio_description,
                                    selected_symbols=st.session_state.portfolio_selected,
                                    shares_data=st.session_state.portfolio_shares,
                                    tax_rates_data=st.session_state.portfolio_tax_rates
                                )
                                if success:
                                    st.success(f"✅ Portfolio '{portfolio_name}' guardado correctamente")
                                else:
                                    st.error("❌ Error al guardar el portfolio. Verifica que el nombre no esté duplicado.")
                            else:
                                st.warning("⚠️ Por favor, ingresa un nombre para el portfolio")
                    
                    # Tabla resumen
                    st.markdown("---")
                    st.markdown("### 📋 Resumen del Portfolio")
                    summary_df = pd.DataFrame(portfolio_data)
                    st.dataframe(
                        summary_df[['symbol', 'name', 'shares', 'price', 'total_cost', 'annual_dividend']],
                        use_container_width=True
                    )
            else:
                st.info("💡 Selecciona acciones para comenzar a construir tu portfolio")
    
    def _portfolio_calendar_tab(self):
        """Tab para mostrar el calendario de dividendos."""
        st.markdown("### 📅 Calendario Anual de Dividendos")
        
        # Mostrar información del portfolio cargado si existe
        if st.session_state.get('portfolio_loaded_name'):
            st.info(f"📂 Portfolio cargado: **{st.session_state.portfolio_loaded_name}**")
        
        if not st.session_state.portfolio_selected:
            st.warning("⚠️ No hay acciones seleccionadas en tu portfolio.")
            st.info("💡 Ve a la pestaña '📋 Seleccionar Acciones' para agregar acciones o carga un portfolio guardado desde '💾 Portfolios Guardados'")
            return
        
        # Obtener datos de las acciones seleccionadas
        all_assets = self.db.get_all_assets()
        selected_assets = [a for a in all_assets 
                         if a['symbol'] in st.session_state.portfolio_selected]
        
        if not selected_assets:
            st.error("❌ No se encontraron datos para las acciones seleccionadas")
            return
        
        # Calcular dividendos mensuales considerando las cantidades de acciones
        monthly_dividends = self._calculate_monthly_dividends(selected_assets, st.session_state.portfolio_shares)
        
        # Aplicar impuestos a los dividendos mensuales antes de mostrar
        for month in monthly_dividends:
            for asset_info in monthly_dividends[month]['assets']:
                symbol = asset_info['symbol']
                tax_rate = st.session_state.portfolio_tax_rates.get(symbol, 0.0)
                dividend_before_tax = asset_info['amount']
                tax_amount = dividend_before_tax * (tax_rate / 100)
                dividend_after_tax = dividend_before_tax - tax_amount
                
                # Guardar información de bruto, impuesto y neto
                asset_info['amount_before_tax'] = dividend_before_tax
                asset_info['tax_amount'] = tax_amount
                asset_info['amount'] = dividend_after_tax  # El neto
                asset_info['tax_rate'] = tax_rate
            
            # Recalcular total del mes después de impuestos (solo neto)
            monthly_dividends[month]['amount'] = sum(a['amount'] for a in monthly_dividends[month]['assets'])
        
        # Filtrar solo acciones con cantidad > 0
        selected_assets_with_shares = [
            a for a in selected_assets 
            if st.session_state.portfolio_shares.get(a['symbol'], 0) > 0
        ]
        
        if not selected_assets_with_shares:
            st.warning("⚠️ No hay acciones con cantidad asignada.")
            st.info("💡 Ve a la pestaña '📋 Seleccionar Acciones' y asigna cantidades a las acciones")
            return
        
        # Mostrar calendario
        self._display_dividend_calendar(monthly_dividends, selected_assets_with_shares)
    
    def _calculate_monthly_dividends(self, assets: List[Dict], shares_dict: Dict[str, int] = None) -> Dict:
        """
        Calcula los dividendos mensuales basado en la frecuencia de pago y cantidad de acciones.
        
        Args:
            assets: Lista de activos seleccionados
            shares_dict: Diccionario con cantidad de acciones por símbolo {symbol: cantidad}
        
        Returns:
            Diccionario con dividendos por mes
        """
        if shares_dict is None:
            shares_dict = st.session_state.portfolio_shares
        
        monthly_dividends = {i: {'amount': 0.0, 'assets': []} for i in range(1, 13)}
        
        for asset in assets:
            symbol = asset['symbol']
            shares = shares_dict.get(symbol, 0)
            
            if shares <= 0:
                continue
            
            # Dividendo anual por acción
            annual_dividend_per_share = asset.get('annual_dividend', 0) or 0
            # Dividendo anual total (por acción * cantidad de acciones)
            total_annual_dividend = annual_dividend_per_share * shares
            frequency = asset.get('dividend_frequency', 'sin_dividendos')
            
            if total_annual_dividend <= 0 or frequency == 'sin_dividendos':
                continue
            
            # Obtener meses de pago reales desde la BD
            months_data = asset.get('dividend_payment_months', '')
            payment_months = []
            
            # Parsear meses de pago correctamente
            if months_data:
                if isinstance(months_data, str):
                    # Si es string, parsear usando el método de la BD
                    payment_months = self.db._parse_payment_months(months_data)
                elif isinstance(months_data, list):
                    # Si ya es lista, validar y normalizar
                    for m in months_data:
                        if isinstance(m, int) and 1 <= m <= 12:
                            payment_months.append(m)
                        elif isinstance(m, str) and m.strip().isdigit():
                            m_int = int(m.strip())
                            if 1 <= m_int <= 12:
                                payment_months.append(m_int)
                else:
                    # Intentar convertir a string y parsear
                    payment_months = self.db._parse_payment_months(str(months_data))
            
            # Si no hay meses de pago definidos, usar lógica por defecto según frecuencia
            if not payment_months:
                if frequency == 'mensual':
                    payment_months = list(range(1, 13))  # Todos los meses
                elif frequency == 'trimestral':
                    payment_months = [3, 6, 9, 12]  # Meses típicos trimestrales
                elif frequency == 'irregular':
                    payment_months = [6, 12]  # Meses típicos irregulares
            
            # Calcular dividendos mensuales según frecuencia y meses reales
            if frequency == 'mensual':
                # Dividendo mensual = anual / número de meses que paga
                num_payments = len(payment_months) if payment_months else 12
                monthly_amount = total_annual_dividend / num_payments
                # Distribuir en los meses que realmente paga
                for month in payment_months:
                    monthly_dividends[month]['amount'] += monthly_amount
                    monthly_dividends[month]['assets'].append({
                        'symbol': symbol,
                        'name': asset.get('name', symbol),
                        'amount': monthly_amount,
                        'shares': shares
                    })
            
            elif frequency == 'trimestral':
                # Dividendo trimestral = anual / número de pagos trimestrales
                num_payments = len(payment_months) if payment_months else 4
                quarterly_amount = total_annual_dividend / num_payments
                # Distribuir en los meses reales de pago
                for month in payment_months:
                    monthly_dividends[month]['amount'] += quarterly_amount
                    monthly_dividends[month]['assets'].append({
                        'symbol': symbol,
                        'name': asset.get('name', symbol),
                        'amount': quarterly_amount,
                        'shares': shares
                    })
            
            elif frequency == 'irregular':
                # Dividendo irregular = anual / número de pagos
                num_payments = len(payment_months) if payment_months else 2
                irregular_amount = total_annual_dividend / num_payments
                # Distribuir en los meses reales de pago
                for month in payment_months:
                    monthly_dividends[month]['amount'] += irregular_amount
                    monthly_dividends[month]['assets'].append({
                        'symbol': symbol,
                        'name': asset.get('name', symbol),
                        'amount': irregular_amount,
                        'shares': shares
                    })
        
        return monthly_dividends
    
    def _display_dividend_calendar(self, monthly_dividends: Dict, assets: List[Dict]):
        """Muestra el calendario de dividendos."""
        month_names = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        
        # Calcular total anual
        total_annual = sum(data['amount'] for data in monthly_dividends.values())
        
        # Mostrar resumen anual
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Anual Proyectado", format_currency(total_annual))
        with col2:
            avg_monthly = total_annual / 12
            st.metric("Promedio Mensual", format_currency(avg_monthly))
        with col3:
            # Calcular total de acciones
            total_shares = sum(st.session_state.portfolio_shares.get(a['symbol'], 0) for a in assets)
            st.metric("Total de Acciones", f"{int(total_shares)}")
        
        st.markdown("---")
        
        # Mostrar calendario por trimestres
        quarters = [
            ([1, 2, 3], "Q1 - Primer Trimestre"),
            ([4, 5, 6], "Q2 - Segundo Trimestre"),
            ([7, 8, 9], "Q3 - Tercer Trimestre"),
            ([10, 11, 12], "Q4 - Cuarto Trimestre")
        ]
        
        for quarter_months, quarter_title in quarters:
            st.markdown(f"### {quarter_title}")
            
            cols = st.columns(3)
            for idx, month in enumerate(quarter_months):
                with cols[idx]:
                    month_data = monthly_dividends[month]
                    st.markdown(f"#### {month_names[month]}")
                    
                    if month_data['amount'] > 0:
                        # Mostrar total neto del mes
                        st.success(f"**Total Neto:** {format_currency(month_data['amount'])}")
                        
                        # Mostrar acciones que pagan con desglose de impuestos
                        if month_data['assets']:
                            st.markdown("**Acciones:**")
                            for asset_info in month_data['assets']:
                                symbol = asset_info['symbol']
                                shares_info = f" ({asset_info.get('shares', 0)} acciones)" if asset_info.get('shares') else ""
                                
                                # Obtener valores (pueden venir con o sin impuestos aplicados)
                                amount_before_tax = asset_info.get('amount_before_tax', asset_info.get('amount', 0))
                                tax_amount = asset_info.get('tax_amount', 0)
                                amount_after_tax = asset_info.get('amount', amount_before_tax - tax_amount)
                                tax_rate = asset_info.get('tax_rate', 0.0)
                                
                                # Si hay impuesto, mostrar desglose
                                if tax_amount > 0:
                                    st.write(
                                        f"• **{symbol}**{shares_info}: "
                                        f"{format_currency(amount_after_tax)} "
                                        f"({format_currency(amount_before_tax)} - "
                                        f"{format_currency(tax_amount)} imp. {tax_rate:.1f}%)"
                                    )
                                else:
                                    st.write(f"• **{symbol}**{shares_info}: {format_currency(amount_after_tax)}")
                    else:
                        st.info("Sin dividendos este mes")
            
            st.markdown("---")
        
        # Gráfico de barras mensual
        st.markdown("### 📊 Visualización Mensual")
        
        months = [month_names[i] for i in range(1, 13)]
        amounts = [monthly_dividends[i]['amount'] for i in range(1, 13)]
        
        chart_data = pd.DataFrame({
            'Mes': months,
            'Dividendos (USD)': amounts
        })
        
        st.bar_chart(chart_data.set_index('Mes'))
    
    def _portfolio_summary_tab(self):
        """Tab para mostrar resumen financiero del portfolio."""
        st.markdown("### 💰 Resumen Financiero del Portfolio")
        
        # Mostrar información del portfolio cargado si existe
        if st.session_state.get('portfolio_loaded_name'):
            st.info(f"📂 Portfolio cargado: **{st.session_state.portfolio_loaded_name}**")
        
        if not st.session_state.portfolio_selected:
            st.warning("⚠️ No hay acciones seleccionadas en tu portfolio.")
            st.info("💡 Ve a la pestaña '📋 Seleccionar Acciones' para agregar acciones o carga un portfolio guardado desde '💾 Portfolios Guardados'")
            return
        
        # Obtener datos
        all_assets = self.db.get_all_assets()
        selected_assets = [a for a in all_assets 
                         if a['symbol'] in st.session_state.portfolio_selected]
        
        if not selected_assets:
            st.error("❌ No se encontraron datos")
            return
        
        # Calcular métricas considerando las cantidades de acciones
        shares_dict = st.session_state.portfolio_shares
        
        # Filtrar solo acciones con cantidad > 0
        selected_assets_with_shares = [
            a for a in selected_assets 
            if shares_dict.get(a['symbol'], 0) > 0
        ]
        
        if not selected_assets_with_shares:
            st.warning("⚠️ No hay acciones con cantidad asignada.")
            st.info("💡 Ve a la pestaña '📋 Seleccionar Acciones' y asigna cantidades")
            return
        
        # Calcular costos y dividendos con cantidades reales (considerando impuestos)
        total_cost = 0
        total_annual_dividend_before_tax = 0
        total_annual_dividend_after_tax = 0
        total_tax_amount = 0
        weighted_yield_sum = 0
        total_investment = 0
        
        for asset in selected_assets_with_shares:
            symbol = asset['symbol']
            shares = shares_dict.get(symbol, 0)
            price = asset.get('current_price', 0)
            annual_div_per_share = asset.get('annual_dividend', 0) or 0
            tax_rate = st.session_state.portfolio_tax_rates.get(symbol, 0.0)
            
            asset_cost = price * shares
            asset_annual_dividend_before_tax = annual_div_per_share * shares
            asset_tax_amount = asset_annual_dividend_before_tax * (tax_rate / 100)
            asset_annual_dividend_after_tax = asset_annual_dividend_before_tax - asset_tax_amount
            
            total_cost += asset_cost
            total_annual_dividend_before_tax += asset_annual_dividend_before_tax
            total_annual_dividend_after_tax += asset_annual_dividend_after_tax
            total_tax_amount += asset_tax_amount
            total_investment += asset_cost
        
        # Calcular yield promedio ponderado (después de impuestos)
        if total_investment > 0:
            avg_yield = (total_annual_dividend_after_tax / total_investment) * 100
        else:
            avg_yield = 0
        
        # Calcular dividendos mensuales (después de impuestos)
        monthly_dividends = self._calculate_monthly_dividends(selected_assets_with_shares, shares_dict)
        # Aplicar impuestos a los dividendos mensuales
        for month in monthly_dividends:
            for asset_info in monthly_dividends[month]['assets']:
                symbol = asset_info['symbol']
                tax_rate = st.session_state.portfolio_tax_rates.get(symbol, 0.0)
                dividend_before_tax = asset_info['amount']
                tax_amount = dividend_before_tax * (tax_rate / 100)
                asset_info['amount'] = dividend_before_tax - tax_amount
                asset_info['tax_amount'] = tax_amount
            
            # Recalcular total del mes después de impuestos
            monthly_dividends[month]['amount'] = sum(a['amount'] for a in monthly_dividends[month]['assets'])
        
        total_monthly = sum(data['amount'] for data in monthly_dividends.values())
        avg_monthly = total_monthly / 12
        
        # Sección destacada: Ganancias e Impuestos
        st.markdown("### 💰 Ganancias e Impuestos")
        st.markdown("---")
        
        # Métricas principales de ganancias
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("#### 💵 Ganancias Brutas (Antes de Impuestos)")
            st.metric("Dividendos Anuales Totales", format_currency(total_annual_dividend_before_tax), 
                     delta=None, delta_color="normal")
            st.caption(f"Promedio mensual: {format_currency(total_annual_dividend_before_tax / 12)}")
        
        with col2:
            st.markdown("#### 💸 Impuestos Totales")
            tax_percentage = (total_tax_amount / total_annual_dividend_before_tax * 100) if total_annual_dividend_before_tax > 0 else 0
            st.metric("Total Descontado", format_currency(total_tax_amount), 
                     delta=f"-{tax_percentage:.1f}%", delta_color="inverse")
            st.caption(f"Promedio mensual: {format_currency(total_tax_amount / 12)}")
        
        with col3:
            st.markdown("#### ✅ Ganancias Netas (Después de Impuestos)")
            st.metric("**Lo que realmente ganas**", format_currency(total_annual_dividend_after_tax), 
                     delta=f"{format_currency(total_annual_dividend_after_tax - total_tax_amount)}", delta_color="normal")
            st.caption(f"Promedio mensual: {format_currency(avg_monthly)}")
        
        st.markdown("---")
        
        # Visualización de desglose
        st.markdown("#### 📊 Desglose Visual")
        col1, col2, col3 = st.columns([2, 1, 2])
        
        with col1:
            st.markdown(f"""
            <div style="background-color: #e8f5e9; padding: 15px; border-radius: 10px; border-left: 5px solid #4caf50;">
                <h4 style="color: #2e7d32; margin: 0;">💰 Ganancias Brutas</h4>
                <p style="font-size: 24px; font-weight: bold; color: #1b5e20; margin: 10px 0;">
                    {format_currency(total_annual_dividend_before_tax)}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div style="text-align: center; padding: 15px;">
                <p style="font-size: 20px; margin: 0;">➖</p>
                <p style="font-size: 18px; font-weight: bold; color: #d32f2f; margin: 5px 0;">
                    {format_currency(total_tax_amount)}
                </p>
                <p style="font-size: 14px; color: #666; margin: 0;">Impuestos</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div style="background-color: #fff3e0; padding: 15px; border-radius: 10px; border-left: 5px solid #ff9800;">
                <h4 style="color: #e65100; margin: 0;">✅ Ganancias Netas</h4>
                <p style="font-size: 24px; font-weight: bold; color: #bf360c; margin: 10px 0;">
                    {format_currency(total_annual_dividend_after_tax)}
                </p>
                <p style="font-size: 14px; color: #666; margin: 0;">Lo que realmente recibes</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Métricas adicionales
        st.markdown("### 📈 Métricas del Portfolio")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("💰 Costo Total del Portfolio", format_currency(total_cost))
        
        with col2:
            st.metric("📈 Yield del Portfolio (neto)", format_percentage(avg_yield))
        
        with col3:
            # Calcular total de acciones
            total_shares = sum(shares_dict.get(a['symbol'], 0) for a in selected_assets_with_shares)
            st.metric("📊 Total de Acciones", f"{int(total_shares)}")
        
        with col4:
            if total_shares > 0:
                avg_price = total_cost / total_shares
                st.metric("💲 Precio Promedio por Acción", format_currency(avg_price))
        
        # ROI (después de impuestos)
        if total_cost > 0:
            roi = (total_annual_dividend_after_tax / total_cost) * 100
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("🎯 ROI Anual Estimado (neto)", format_percentage(roi))
            with col2:
                # Calcular porcentaje de impuestos sobre ganancias brutas
                tax_percentage_of_gross = (total_tax_amount / total_annual_dividend_before_tax * 100) if total_annual_dividend_before_tax > 0 else 0
                st.metric("📉 Impuestos sobre Ganancias", f"{tax_percentage_of_gross:.2f}%")
        
        # Tabla detallada con cantidades
        st.markdown("---")
        st.markdown("### 📋 Detalle del Portfolio")
        
        portfolio_detail = []
        month_names_short = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }
        
        for asset in selected_assets_with_shares:
            symbol = asset['symbol']
            shares = shares_dict.get(symbol, 0)
            price = asset.get('current_price', 0)
            annual_div_per_share = asset.get('annual_dividend', 0) or 0
            tax_rate = st.session_state.portfolio_tax_rates.get(symbol, 0.0)
            
            # Calcular dividendos después de impuestos
            annual_dividend_total_before_tax = annual_div_per_share * shares
            tax_amount = annual_dividend_total_before_tax * (tax_rate / 100)
            annual_dividend_total_after_tax = annual_dividend_total_before_tax - tax_amount
            
            # Obtener y formatear meses de pago
            months_data = asset.get('dividend_payment_months', '')
            payment_months_str = 'N/A'
            
            if months_data:
                payment_months = []
                if isinstance(months_data, str):
                    payment_months = self.db._parse_payment_months(months_data)
                elif isinstance(months_data, list):
                    for m in months_data:
                        if isinstance(m, int) and 1 <= m <= 12:
                            payment_months.append(m)
                        elif isinstance(m, str) and m.strip().isdigit():
                            m_int = int(m.strip())
                            if 1 <= m_int <= 12:
                                payment_months.append(m_int)
                else:
                    payment_months = self.db._parse_payment_months(str(months_data))
                
                if payment_months:
                    month_labels = [month_names_short.get(m, str(m)) for m in sorted(payment_months)]
                    payment_months_str = ', '.join(month_labels)
            
            portfolio_detail.append({
                'Símbolo': symbol,
                'Nombre': asset.get('name', 'N/A'),
                'Cantidad': shares,
                'Precio Unitario': format_currency(price),
                'Costo Total': format_currency(price * shares),
                'Dividendo Anual (por acción)': format_currency(annual_div_per_share),
                'Dividendo Anual Total (bruto)': format_currency(annual_dividend_total_before_tax),
                'Impuesto (%)': f"{tax_rate:.1f}%",
                'Impuesto (USD)': format_currency(tax_amount),
                'Dividendo Anual Total (neto)': format_currency(annual_dividend_total_after_tax),
                'Yield': format_percentage(asset.get('dividend_yield', 0)),
                'Frecuencia': asset.get('dividend_frequency', 'N/A'),
                'Meses de Pago': payment_months_str
            })
        
        detail_df = pd.DataFrame(portfolio_detail)
        st.dataframe(detail_df, use_container_width=True)
    
    def _portfolio_saved_tab(self):
        """Tab para gestionar portfolios guardados."""
        st.markdown("### 💾 Portfolios Guardados")
        
        # Obtener todos los portfolios
        portfolios = self.db.get_all_portfolios()
        
        if not portfolios:
            st.info("💡 No hay portfolios guardados. Ve a la pestaña '📋 Seleccionar Acciones' para crear y guardar uno.")
            return
        
        st.write(f"**Total de portfolios guardados:** {len(portfolios)}")
        st.markdown("---")
        
        # Mostrar lista de portfolios
        for portfolio in portfolios:
            with st.expander(f"📁 {portfolio['name']} - Actualizado: {portfolio['updated_at'][:10] if portfolio['updated_at'] else 'N/A'}"):
                col1, col2, col3 = st.columns([3, 1, 1])
                
                with col1:
                    if portfolio.get('description'):
                        st.write(f"**Descripción:** {portfolio['description']}")
                    st.write(f"**Acciones:** {len(portfolio['selected_symbols'])}")
                    st.write(f"**Símbolos:** {', '.join(portfolio['selected_symbols'][:10])}")
                    if len(portfolio['selected_symbols']) > 10:
                        st.caption(f"... y {len(portfolio['selected_symbols']) - 10} más")
                    st.caption(f"Creado: {portfolio['created_at'][:19] if portfolio.get('created_at') else 'N/A'}")
                
                with col2:
                    if st.button("📂 Cargar", key=f"load_{portfolio['name']}", use_container_width=True):
                        try:
                            # Cargar portfolio en session_state
                            # Asegurar que los datos sean del tipo correcto
                            selected_symbols = portfolio.get('selected_symbols', [])
                            shares_data = portfolio.get('shares_data', {})
                            tax_rates_data = portfolio.get('tax_rates_data', {})
                            
                            # Convertir a listas/diccionarios si vienen como strings
                            if isinstance(selected_symbols, str):
                                import json
                                selected_symbols = json.loads(selected_symbols)
                            if isinstance(shares_data, str):
                                import json
                                shares_data = json.loads(shares_data)
                            if isinstance(tax_rates_data, str):
                                import json
                                tax_rates_data = json.loads(tax_rates_data)
                            
                            # Asegurar que shares_data y tax_rates_data sean diccionarios con valores numéricos
                            shares_data_clean = {}
                            for k, v in shares_data.items():
                                try:
                                    shares_data_clean[str(k)] = int(v) if isinstance(v, (int, float)) else 0
                                except:
                                    shares_data_clean[str(k)] = 0
                            
                            tax_rates_data_clean = {}
                            for k, v in tax_rates_data.items():
                                try:
                                    tax_rates_data_clean[str(k)] = float(v) if isinstance(v, (int, float)) else 0.0
                                except:
                                    tax_rates_data_clean[str(k)] = 0.0
                            
                            # Guardar en session_state - normalizar todos los símbolos a mayúsculas para consistencia
                            portfolio_selected_clean = [str(s).upper().strip() for s in selected_symbols if s] if selected_symbols else []
                            portfolio_shares_clean = {str(k).upper().strip(): v for k, v in shares_data_clean.items()}
                            portfolio_tax_rates_clean = {str(k).upper().strip(): v for k, v in tax_rates_data_clean.items()}
                            
                            st.session_state.portfolio_selected = portfolio_selected_clean
                            st.session_state.portfolio_shares = portfolio_shares_clean
                            st.session_state.portfolio_tax_rates = portfolio_tax_rates_clean
                            st.session_state.portfolio_loaded_name = portfolio['name']
                            
                            logger.info(f"Portfolio '{portfolio['name']}' cargado: {len(st.session_state.portfolio_selected)} acciones, {len(st.session_state.portfolio_shares)} con shares")
                            logger.info(f"Shares cargados: {st.session_state.portfolio_shares}")
                            logger.info(f"Tax rates cargados: {st.session_state.portfolio_tax_rates}")
                            
                            st.success(f"✅ Portfolio '{portfolio['name']}' cargado correctamente. Ve a la pestaña '📋 Seleccionar Acciones' para verlo.")
                            st.rerun()
                        except Exception as e:
                            logger.error(f"Error cargando portfolio: {e}", exc_info=True)
                            st.error(f"❌ Error al cargar el portfolio: {e}")
                
                with col3:
                    if st.button("🗑️ Eliminar", key=f"delete_{portfolio['name']}", use_container_width=True):
                        if self.db.delete_portfolio(portfolio['name']):
                            st.success(f"✅ Portfolio '{portfolio['name']}' eliminado")
                            st.rerun()
                        else:
                            st.error("❌ Error al eliminar el portfolio")
        
        st.markdown("---")
        st.markdown("### 📊 Resumen de Portfolios")
        
        # Crear DataFrame con resumen
        summary_data = []
        for portfolio in portfolios:
            total_shares = sum(portfolio['shares_data'].values())
            total_actions = len(portfolio['selected_symbols'])
            summary_data.append({
                'Nombre': portfolio['name'],
                'Acciones': total_actions,
                'Total Shares': total_shares,
                'Última Actualización': portfolio['updated_at'][:10] if portfolio.get('updated_at') else 'N/A'
            })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True)
    
    def run(self):
        """Método principal que ejecuta la aplicación."""
        try:
            self.render_header()
            page = self.render_sidebar()
            
            if page == "🏠 Inicio":
                st.info("👈 Selecciona una opción en el sidebar para comenzar")
                st.markdown("""
                ### 🚀 Comienza aquí:
                
                1. **📥 Importar Excel**: Sube un archivo con múltiples tickers
                2. **🔍 Buscar Activo**: Analiza un activo individual
                3. **📊 Ver Activos**: Explora los activos guardados
                4. **📈 Visualizaciones**: Descubre patrones y oportunidades
                5. **ℹ️ Estadísticas**: Revisa métricas agregadas
                """)
            
            elif page == "📥 Importar Excel":
                self.import_excel_page()
            
            elif page == "🔍 Buscar Activo":
                self.search_asset_page()
            
            elif page == "📊 Ver Activos":
                self.view_assets_page()
            
            elif page == "📈 Visualizaciones":
                self.visualizations_page()
            
            elif page == "ℹ️ Estadísticas":
                self.stats_page()
            
            elif page == "🔧 Mantenimiento":
                self.maintenance_page()
            
            elif page == "🏪 Gestión de Plataformas":
                self.platforms_management_page()
            
            elif page == "💼 Constructor de Portfolio":
                self.portfolio_builder_page()
                
        except Exception as e:
            logger.error(f"Error en ejecución: {e}", exc_info=True)
            st.error("❌ Error en la aplicación. Por favor, revisa los logs.")


# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":
    """
    Para ejecutar la aplicación:
    
    streamlit run app.py
    """
    try:
        app = DividendHunterApp()
        app.run()
    except Exception as e:
        st.error(f"❌ Error fatal: {e}")
        logger.critical(f"Error fatal en aplicación: {e}", exc_info=True)

